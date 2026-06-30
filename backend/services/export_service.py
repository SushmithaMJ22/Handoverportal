"""
ExportService — produces CSV, Excel, and PDF exports of HandoverRecord lists.

All three formats match the column set and styling used in the original
reports.py inline implementations, now consolidated here.
"""
import csv
import io
import logging
from datetime import datetime
from typing import List

from models import HandoverRecord

logger = logging.getLogger(__name__)

# Shared column definitions (label → attribute name on HandoverRecord)
_COLUMNS = [
    ("ID", "id"),
    ("Customer Name", "customer_name"),
    ("Contact Person", "contact_person"),
    ("Phone", "contact_phone"),
    ("Email", "contact_email"),
    ("Region", "region"),
    ("Product", "product"),
    ("Sub Product", "sub_product"),
    ("Platform", "platform"),
    ("Solution", "solution"),
    ("PS Engineer", "ps_engineer"),
    ("Sales Person", "sales_person"),
    ("PS Reviewer", "ps_reviewer"),
    ("Support Ticket", "support_ticket"),
    ("Support Reviewer", "support_reviewer"),
    ("Remarks", "remarks"),
    ("Status", "status"),
    ("Created At", "created_at"),
]

_FALLBACK = {
    "sub_product": "N/A",
    "solution": "N/A",
    "status": "active",
}


def _get_value(record: HandoverRecord, attr: str) -> str:
    """Extract a display-safe string value from a HandoverRecord field."""
    val = getattr(record, attr, None)
    if val is None:
        return _FALLBACK.get(attr, "")
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    return str(val)


class ExportService:
    def to_csv(self, records: List[HandoverRecord]) -> io.BytesIO:
        """Return a UTF-8-with-BOM CSV as a BytesIO buffer."""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([label for label, _ in _COLUMNS])
        for r in records:
            writer.writerow([_get_value(r, attr) for _, attr in _COLUMNS])

        output.seek(0)
        return io.BytesIO(output.getvalue().encode("utf-8-sig"))

    def to_xlsx(self, records: List[HandoverRecord]) -> io.BytesIO:
        """Return a styled Excel workbook as a BytesIO buffer."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Handover Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )

        for col_idx, (label, _) in enumerate(_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = 18

        for row_idx, r in enumerate(records, 2):
            for col_idx, (_, attr) in enumerate(_COLUMNS, 1):
                ws.cell(row=row_idx, column=col_idx, value=_get_value(r, attr))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def to_pdf(self, records: List[HandoverRecord]) -> io.BytesIO:
        """Return a landscape A4 PDF table as a BytesIO buffer."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        # PDF uses a condensed column set to fit landscape A4
        pdf_columns = [
            ("Customer", "customer_name"),
            ("Region", "region"),
            ("Product", "product"),
            ("Platform", "platform"),
            ("Engineer", "ps_engineer"),
            ("Ticket", "support_ticket"),
            ("Status", "status"),
            ("Date", "created_at"),
        ]

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Handover Report", styles["Title"]))
        elements.append(
            Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                styles["Normal"],
            )
        )
        elements.append(Spacer(1, 20))

        data = [[label for label, _ in pdf_columns]]
        for r in records:
            row = []
            for _, attr in pdf_columns:
                val = getattr(r, attr, None)
                if isinstance(val, datetime):
                    row.append(val.strftime("%Y-%m-%d"))
                else:
                    row.append(str(val) if val else "")
            data.append(row)

        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F3F4F6")],
                    ),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output


export_service = ExportService()
